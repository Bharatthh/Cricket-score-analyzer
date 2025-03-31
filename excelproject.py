from openpyxl import Workbook
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
ch=input("if you want to add score press y: ")
while ch=="y":
    wb=Workbook()
    sheet=wb.active
    n="y"
    i=2
    sheet["A1"]="over"
    sheet["B1"]="TEAM1"
    sheet["C1"]="TEAM2"

    while n=="y":
        sheet[f"A{i}"]=int(input("enter no of over : "))
        sheet[f"B{i}"]=int(input("enter the team1 score : "))
        sheet[f"C{i}"] = int(input("enter the team2 score : "))
        i = i + 1
        n=input("enter y for continue: ")
    wb.save(filename="excelproject.xlsx")
    ch="n"
path="C:\\Users\\user\\pythonProject\\excelproject.xlsx"
wb_file=openpyxl.load_workbook(path)
sheet=wb_file.active
max_col=sheet.max_column
max_row_=sheet.max_row
over=[]
team1=[]
team1_wic=[]
team2=[]
team2_wic=[]
print(max_col)
count=1
for i in range(1,max_col+1):
    for j in range(2,max_row_+1):
        cellval = sheet.cell(row=j, column=i)
        if i==1:
            over.append(cellval.value)
        elif i==2:
            team1.append(cellval.value)
        elif i==3:
            team1_wic.append(cellval.value)
        elif i==4:
            team2.append(cellval.value)
        elif i==5:
            team2_wic.append(cellval.value)
    print()

print(over)
print(team1)
print(team2)

# df=pd.DataFrame({1:"team1",2:"team2",3:"over"})
df = pd.DataFrame({
    'Over': over,
    'Team1': team1,
    'team1_wic':team1_wic,
    'Team2': team2,
    'team2_wic':team2_wic
})
print(df)
ax1=df.plot(x='Over', y=['Team1', 'Team2'])
ax1.scatter(df['Over'][df['team1_wic'] > 0], df['Team1'][df['team1_wic'] > 0], color='red', label='Team1 Wicket', marker='o')
ax1.scatter(df['Over'][df['team2_wic'] > 0], df['Team2'][df['team2_wic'] > 0], color='yellow', label='Team2 Wicket', marker='o')

plt.show()


